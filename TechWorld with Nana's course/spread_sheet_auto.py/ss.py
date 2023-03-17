import openpyxl

inv_file = openpyxl.load_workbook("C:\\Users\\Eyob 2\\Desktop\\PYTHON\\TechWorld with Nana's course\\spread_sheet_auto.py\\inventory.xlsx")

product_list = inv_file["Sheet1"]

products_per_supplier = {}

print(product_list.max_row)

for row in range (2, product_list.max_row+1):
    supplier_name = product_list.cell(row, 4)

    for i,j in products_per_supplier:

        if supplier_name not in i:
            products_per_supplier[supplier_name] = 1
        
        else:
            products_per_supplier[supplier_name] = j+1

print(products_per_supplier)
